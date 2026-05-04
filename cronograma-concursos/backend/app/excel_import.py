"""Excel bulk import for the Cronograma Concurso workbook.

Workbook layout (one sheet per concurso):

    Row 0: optional title row. Column B holds the canonical project name
           (e.g. "Huawei Innovation Competition"). Preferred source for
           the project name; falls back to the first `Concurso` cell, then
           the sheet tab name.
    Row 1: header row starting with `Concurso` in column A, then
           `Secuencia / Pasos`, `Fecha de inicio`, `Fecha de fin`,
           `Duración (días)`, `Completo`, `Notas`. Legacy columns
           (`Responsable`, `Observaciones`, `Status`) are still accepted
           on input.
    Row 2+: data rows. Column A typically repeats the project name; tasks
            live in the remaining columns.
    After the task list, a metadata block — rows whose column A is one of:
           `Contacto`           → captured into `Project.contact_name`
           `Ficha de proyecto`  → captured into `Project.description`
                                  (used as a URL link in the UI)

Deduplication
-------------
Projects are matched by exact name. Tasks are upserted by the natural key
`(project_id, lower(name))` so re-uploading the same workbook updates
existing rows instead of duplicating them. With `replace_tasks=True`, the
project's existing tasks are deleted first so the sheet becomes the
source of truth.
"""
from __future__ import annotations

import io
import logging
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from . import models, schemas

logger = logging.getLogger(__name__)

# Excel's epoch for date serials (the "1900" system, with the leap-year bug).
_EXCEL_EPOCH = date(1899, 12, 30)

COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "project_name": ("concurso",),
    "name": ("secuencia / pasos", "secuencia/pasos", "secuencia", "pasos", "tarea"),
    "start_date": ("fecha de inicio", "inicio"),
    "end_date": ("fecha de fin", "fin", "fecha fin"),
    "duration_days": ("duracion (dias)", "duración (días)", "duracion", "duración"),
    "complete": ("completo", "completado"),
    # Legacy: still parsed for backward compatibility with older workbooks.
    "responsible": ("responsable", "encargado"),
    "observations": ("notas", "observaciones"),
}

# Metadata rows after the task list. Column A label → project field.
_METADATA_LABELS: dict[str, str] = {
    "contacto": "contact_name",
    "ficha de proyecto": "description",
    "ficha": "description",
}

TEMPLATE_HEADERS: tuple[str, ...] = (
    "Concurso",
    "Secuencia / Pasos",
    "Fecha de inicio",
    "Fecha de fin",
    "Duración (días)",
    "Completo",
    "Notas",
)


@dataclass
class _Counters:
    projects_created: int = 0
    projects_updated: int = 0
    tasks_created: int = 0
    tasks_updated: int = 0
    tasks_deleted: int = 0
    skipped_rows: int = 0
    errors: list[str] = field(default_factory=list)


def _normalise(text: object) -> str:
    if text is None:
        return ""
    return str(text).strip().lower()


def _match_header(header_cells: Iterable[object]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(header_cells):
        label = _normalise(cell)
        if not label:
            continue
        for ffield, aliases in COLUMN_ALIASES.items():
            if ffield in mapping:
                continue
            if any(alias in label for alias in aliases):
                mapping[ffield] = idx
                break
    return mapping


def _to_date(value: object) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return _EXCEL_EPOCH + timedelta(days=int(value))
        except (OverflowError, ValueError):
            return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    return None


def _to_bool(value: object) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    text = _normalise(value)
    if text in {"si", "sí", "yes", "true", "1", "✅", "x", "completo", "completado"}:
        return True
    return False


def _to_int(value: object) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _clean_text(value: object) -> str | None:
    if value is None:
        return None
    text = " ".join(str(value).split())
    return text or None


def _looks_like_section_header(name: str, other_cells_empty: bool) -> bool:
    stripped = name.strip()
    if not stripped:
        return True
    if not other_cells_empty:
        return False
    letters = [c for c in stripped if c.isalpha()]
    if letters and all(c.isupper() for c in letters) and len(stripped.split()) <= 4:
        return True
    return False


def _first_non_empty(row: Iterable[object]) -> str | None:
    for cell in row:
        text = _clean_text(cell)
        if text:
            return text
    return None


def _find_project_name(rows: list[tuple], header_idx: int, header_map: dict[str, int], fallback: str) -> str:
    """Title row B > first Concurso value > sheet tab name."""
    for r in rows[:header_idx]:
        if len(r) >= 2:
            title = _clean_text(r[1])
            if title:
                return title
        lone = _first_non_empty(r)
        if lone:
            return lone

    cidx = header_map.get("project_name")
    if cidx is not None:
        for r in rows[header_idx + 1 :]:
            if cidx < len(r):
                value = _clean_text(r[cidx])
                if value and value.lower() not in _METADATA_LABELS:
                    return value

    return fallback


def import_workbook(
    db: Session,
    file_bytes: bytes,
    *,
    replace_tasks: bool = False,
) -> schemas.ImportExcelResult:
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        return schemas.ImportExcelResult(
            projects_created=0,
            tasks_created=0,
            skipped_rows=0,
            errors=[f"Archivo inválido: {exc}"],
        )

    counters = _Counters()

    for sheet in wb.worksheets:
        _import_sheet(db, sheet, counters, replace_tasks=replace_tasks)

    db.commit()
    logger.info(
        "excel.import projects_created=%s projects_updated=%s tasks_created=%s "
        "tasks_updated=%s tasks_deleted=%s skipped=%s errors=%s",
        counters.projects_created,
        counters.projects_updated,
        counters.tasks_created,
        counters.tasks_updated,
        counters.tasks_deleted,
        counters.skipped_rows,
        len(counters.errors),
    )
    return schemas.ImportExcelResult(
        projects_created=counters.projects_created,
        projects_updated=counters.projects_updated,
        tasks_created=counters.tasks_created,
        tasks_updated=counters.tasks_updated,
        tasks_deleted=counters.tasks_deleted,
        skipped_rows=counters.skipped_rows,
        errors=counters.errors,
    )


def _import_sheet(
    db: Session,
    sheet,
    counters: _Counters,
    *,
    replace_tasks: bool,
) -> None:
    sheet_label = sheet.title.strip() or "(sin título)"
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return

    # A real header row has the task-name column AND at least one supporting
    # column. This avoids matching prose that happens to contain an alias
    # substring (e.g. instructions text mentioning "Secuencia / Pasos").
    _SUPPORTING = ("start_date", "end_date", "duration_days", "complete", "observations")
    header_idx = -1
    header_map: dict[str, int] = {}
    for idx, row in enumerate(rows):
        candidate = _match_header(row)
        if "name" in candidate and any(f in candidate for f in _SUPPORTING):
            header_map = candidate
            header_idx = idx
            break

    if "name" not in header_map:
        # Documentation/instructions sheets are skipped silently. Other
        # sheets without a recognisable header get reported.
        if sheet_label.strip().lower() not in {"instrucciones", "instructions", "guía", "guia"}:
            counters.errors.append(
                f"Hoja '{sheet_label}': no se encontró columna de tareas."
            )
        return

    project_name = _find_project_name(rows, header_idx, header_map, fallback=sheet_label)

    existing = db.query(models.Project).filter(models.Project.name == project_name).first()
    if existing:
        project = existing
        counters.projects_updated += 1
    else:
        project = models.Project(name=project_name)
        db.add(project)
        db.flush()
        counters.projects_created += 1

    if replace_tasks and existing:
        deleted = (
            db.query(models.Task)
            .filter(models.Task.project_id == project.id)
            .delete(synchronize_session=False)
        )
        counters.tasks_deleted += deleted
        db.flush()

    def cell_for(row: tuple, ffield: str):
        i = header_map.get(ffield)
        if i is None or i >= len(row):
            return None
        return row[i]

    # Build a name → existing Task index for upsert. After replace_tasks=True,
    # this is empty. We also update it as we create rows so duplicates within
    # the same sheet collapse to a single task.
    existing_tasks: dict[str, models.Task] = {}
    if not (replace_tasks and existing):
        for t in db.query(models.Task).filter(models.Task.project_id == project.id).all():
            existing_tasks[t.name.strip().lower()] = t

    order = (
        db.query(models.Task)
        .filter(models.Task.project_id == project.id)
        .count()
    )

    contact_captured: str | None = None
    ficha_captured: str | None = None

    for row in rows[header_idx + 1 :]:
        project_cell = _clean_text(cell_for(row, "project_name"))
        name_cell = _clean_text(cell_for(row, "name"))

        # Metadata rows: column A = "Contacto" | "Ficha de proyecto", value in
        # the next non-empty cell on the row (typically column B, but the
        # name-column index varies depending on how the sheet was authored).
        if project_cell:
            label = project_cell.lower()
            if label in _METADATA_LABELS:
                value = name_cell or _first_non_empty(
                    c for i, c in enumerate(row) if i != header_map.get("project_name")
                )
                if value:
                    if _METADATA_LABELS[label] == "contact_name":
                        contact_captured = value
                    elif _METADATA_LABELS[label] == "description":
                        ficha_captured = value
                continue

        name = name_cell or ""
        other_empty = all(
            cell_for(row, f) in (None, "")
            for f in ("start_date", "end_date", "duration_days", "responsible")
        )
        if not name or _looks_like_section_header(name, other_empty):
            counters.skipped_rows += 1
            continue

        try:
            start = _to_date(cell_for(row, "start_date"))
            end = _to_date(cell_for(row, "end_date"))
            if start and end and end < start:
                counters.errors.append(
                    f"'{project_name}' · '{name}': fecha de fin anterior al inicio (omitida)."
                )
                counters.skipped_rows += 1
                continue

            duration_raw = cell_for(row, "duration_days")
            duration = _to_int(duration_raw)
            if duration is None and start and end:
                duration = (end - start).days

            complete_raw = cell_for(row, "complete")
            responsible = _clean_text(cell_for(row, "responsible"))
            observations = _clean_text(cell_for(row, "observations"))

            key = name.strip().lower()
            existing_task = existing_tasks.get(key)
            if existing_task is not None:
                # Upsert: update only fields the sheet actually carries, so
                # blank cells don't wipe values the user edited in the app.
                if start is not None:
                    existing_task.start_date = start
                if end is not None:
                    existing_task.end_date = end
                if duration is not None:
                    existing_task.duration_days = duration
                if complete_raw not in (None, ""):
                    existing_task.complete = _to_bool(complete_raw)
                if responsible is not None:
                    existing_task.responsible = responsible
                if observations is not None:
                    existing_task.observations = observations
                # Re-importing implies the dates are authoritative again.
                existing_task.auto_status = True
                existing_task.name = name[:300]
                counters.tasks_updated += 1
            else:
                task = models.Task(
                    project_id=project.id,
                    name=name[:300],
                    start_date=start,
                    end_date=end,
                    duration_days=duration,
                    complete=_to_bool(complete_raw),
                    responsible=responsible,
                    observations=observations,
                    auto_status=True,
                    order=order,
                )
                db.add(task)
                existing_tasks[key] = task
                counters.tasks_created += 1
                order += 1
        except Exception as exc:
            counters.errors.append(f"'{project_name}' · '{name}': {exc}")
            counters.skipped_rows += 1

    if contact_captured:
        project.contact_name = contact_captured
    if ficha_captured:
        project.description = ficha_captured


# ---------------------------------------------------------------------------
# Template generator
# ---------------------------------------------------------------------------

def generate_template() -> bytes:
    """Build a fresh .xlsx the user can fill in and re-upload."""
    wb = Workbook()

    # --- Instructions sheet ---
    instructions = wb.active
    instructions.title = "Instrucciones"
    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)

    lines: list[tuple[str, Font | None]] = [
        ("Plantilla de carga masiva de concursos", title_font),
        ("", None),
        ("Cada hoja de este libro representa un concurso.", None),
        ("Renombra la pestaña con el nombre del concurso, o escríbelo en la celda B1.", None),
        ("", None),
        ("Columnas (fila 2):", bold),
        ("  · Concurso              — opcional, repite el nombre del concurso.", None),
        ("  · Secuencia / Pasos     — nombre de la tarea (obligatorio).", None),
        ("  · Fecha de inicio       — formato YYYY-MM-DD o DD/MM/YYYY.", None),
        ("  · Fecha de fin          — idem.", None),
        ("  · Duración (días)       — opcional; se calcula si dejas las fechas.", None),
        ("  · Completo              — sí/no, x, 1/0.", None),
        ("  · Notas                 — texto libre.", None),
        ("", None),
        ("Metadatos del concurso (al final de la hoja):", bold),
        ("  · Contacto              | Nombre del contacto", None),
        ("  · Ficha de proyecto     | https://...", None),
        ("", None),
        ("Cómo se evitan duplicados:", bold),
        ("  · Los concursos se identifican por nombre exacto.", None),
        ("  · Las tareas se identifican por (Concurso, Secuencia / Pasos), sin distinguir mayúsculas.", None),
        ("  · Re-subir el mismo archivo actualiza filas existentes; no las duplica.", None),
        ("  · Las celdas vacías no borran datos: sólo se actualiza lo que escribas en el archivo.", None),
        ("  · Marca 'Reemplazar tareas existentes' al subir si quieres que el archivo sea la fuente de verdad y borre lo demás.", None),
    ]
    for idx, (text, font) in enumerate(lines, start=1):
        cell = instructions.cell(row=idx, column=1, value=text)
        if font:
            cell.font = font
    instructions.column_dimensions["A"].width = 100

    # --- Example concurso sheet ---
    ws = wb.create_sheet("Mi Concurso 2026")

    ws.cell(row=1, column=2, value="Mi Concurso 2026").font = title_font

    header_fill = PatternFill("solid", fgColor="EEEEEE")
    for i, h in enumerate(TEMPLATE_HEADERS, start=1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = bold
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left")

    sample_rows = [
        (
            "Mi Concurso 2026",
            "Alta de concurso en CRM",
            date(2026, 1, 15),
            date(2026, 1, 20),
            None,
            "",
            "Coordinar con marketing antes de publicar",
        ),
        (
            "Mi Concurso 2026",
            "Definir cronograma con cliente",
            date(2026, 1, 21),
            date(2026, 1, 28),
            None,
            "",
            "",
        ),
        (
            "Mi Concurso 2026",
            "Lanzamiento",
            date(2026, 2, 1),
            date(2026, 2, 1),
            None,
            "",
            "",
        ),
    ]
    for r in sample_rows:
        ws.append(r)

    # Blank separator + metadata
    ws.append([])
    ws.append(["Contacto", "Juan Pérez"])
    ws.append(["Ficha de proyecto", "https://drive.google.com/..."])

    widths = (22, 34, 14, 14, 14, 12, 38)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=2, column=i).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
